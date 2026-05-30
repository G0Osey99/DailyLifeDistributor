"""Batch-run route: reassembly handshake, per-batch delete, run lifecycle."""
import io
import time

import pytest

from core import auth, media_session as ms, upload_jobs
from core.session_state import session


@pytest.fixture()
def client(temp_db, monkeypatch, tmp_path):
    monkeypatch.setattr(ms, "_TEMP_ROOT", str(tmp_path / "uploads"))
    from blueprints import media
    media._run_lock = ms.PerUserRunLock()
    media._runs.clear()
    session.entries.clear()
    session.selected_dates.clear()
    session.upload_results.clear()
    auth.reset_lockouts()
    auth.set_password("pw")
    import app as flask_app_module
    flask_app_module.app.config["TESTING"] = True
    with flask_app_module.app.test_client() as c:
        c.post("/login", data={"password": "pw"})
        yield c


def _init(client):
    return client.post("/media/run/init", json={}).get_json()["run_id"]


def _new_file(client, run_id):
    return client.post(f"/media/file/new?run_id={run_id}").get_json()["file_id"]


def _complete_file(client, run_id, file_id, data=b"data"):
    return client.post(
        "/media/upload/chunk",
        data={"run_id": run_id, "file_id": file_id, "chunk_index": "0",
              "total_chunks": "1", "data": (io.BytesIO(data), "blob")},
        content_type="multipart/form-data",
    )


def test_batch_run_rejected_if_file_incomplete(client):
    run_id = _init(client)
    file_id = _new_file(client, run_id)  # never upload its chunk
    resp = client.post("/media/batch/run", json={
        "run_id": run_id, "dates": ["2025-05-21"], "platforms": ["youtube_video"],
        "files": {file_id: {"category": "youtube_video", "date": "2025-05-21"}},
    })
    assert resp.status_code == 409


def test_batch_run_rejects_non_dict_files_with_400(client):
    """TYPE-002: `files`/`overrides` are read with `or {}`, which only
    rescues falsy values. A non-empty wrong-type (JSON list) used to slip
    through and AttributeError on .items() → 500 with the run lock still
    held. It must now be a clean 400."""
    run_id = _init(client)
    resp = client.post("/media/batch/run", json={
        "run_id": run_id, "dates": ["2025-05-21"], "platforms": ["youtube_video"],
        "files": ["not", "a", "dict"],
    })
    assert resp.status_code == 400, resp.get_data(as_text=True)
    resp2 = client.post("/media/batch/run", json={
        "run_id": run_id, "dates": ["2025-05-21"], "platforms": ["youtube_video"],
        "files": {}, "overrides": ["bad"],
    })
    assert resp2.status_code == 400


def test_batch_run_happy_path_deletes_temp_files(client, monkeypatch):
    captured = {}

    def fake_run_batch(**kwargs):
        captured["file_paths"] = dict(kwargs["file_paths"])
        return set(kwargs["file_paths"].values())

    monkeypatch.setattr(upload_jobs, "run_batch", fake_run_batch)

    run_id = _init(client)
    file_id = _new_file(client, run_id)
    assert _complete_file(client, run_id, file_id).get_json()["complete"] is True

    from blueprints import media
    temp_path = media._runs[run_id]["dir"].file_path(file_id)
    import os
    assert os.path.isfile(temp_path)

    resp = client.post("/media/batch/run", json={
        "run_id": run_id, "dates": ["2025-05-21"], "platforms": ["youtube_video"],
        "files": {file_id: {"category": "youtube_video", "date": "2025-05-21"}},
    })
    assert resp.status_code == 200
    job_id = resp.get_json()["job_id"]

    # Wait for the worker thread to finish.
    deadline = time.time() + 10
    while time.time() < deadline:
        job = upload_jobs.get_job(job_id)
        if job and job.get("done"):
            break
        time.sleep(0.05)
    assert captured["file_paths"][("youtube_video", "2025-05-21")] == temp_path
    assert not os.path.exists(temp_path)  # batch temp file deleted
    # The run's byte counter dropped back to 0 once the batch was deleted, so
    # the per-run ceiling tracks concurrent (per-batch) disk, not a cumulative
    # total — a later batch in the same run isn't penalized for earlier ones.
    assert media._runs[run_id]["bytes_total"] == 0


def test_batch_entries_carry_spreadsheet_metadata(client, monkeypatch):
    """The batch run must upload with the mapped titles, not blanks."""
    import io
    import openpyxl

    captured = {}

    def fake_run_batch(**kwargs):
        captured["entries"] = kwargs["entries_snapshot"]
        captured["summary"] = kwargs["summary"]
        return set(kwargs["file_paths"].values())

    monkeypatch.setattr(upload_jobs, "run_batch", fake_run_batch)

    # Upload a sheet mapping Date -> Title and persist the column mapping.
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Plan"
    ws.append(["Date", "Title"])
    ws.append(["2025-05-21", "Gratitude Today"])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    client.post("/media/spreadsheet", data={"file": (buf, "p.xlsx")},
                content_type="multipart/form-data")
    client.post("/media/mapping", json={
        "sheet_name": "Plan", "date_column": "Date", "youtube_title_column": "Title",
    })

    run_id = _init(client)
    file_id = _new_file(client, run_id)
    _complete_file(client, run_id, file_id)
    resp = client.post("/media/batch/run", json={
        "run_id": run_id, "dates": ["2025-05-21"], "platforms": ["youtube_video"],
        "files": {file_id: {"category": "youtube_video", "date": "2025-05-21"}},
    })
    assert resp.status_code == 200
    job_id = resp.get_json()["job_id"]
    deadline = time.time() + 10
    while time.time() < deadline and not (upload_jobs.get_job(job_id) or {}).get("done"):
        time.sleep(0.05)

    entry = captured["entries"]["2025-05-21"]
    assert entry.youtube_title == "Gratitude Today"   # not blank
    assert entry.youtube_video_path  # points at the uploaded temp file
    assert entry.platforms_enabled.get("youtube_video") is True


def test_batch_overrides_win_over_spreadsheet(client, monkeypatch):
    """A per-date customize override replaces the spreadsheet value."""
    import io
    import openpyxl

    captured = {}
    monkeypatch.setattr(upload_jobs, "run_batch",
                        lambda **k: (captured.update(entries=k["entries_snapshot"]) or
                                     set(k["file_paths"].values())))

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Plan"
    ws.append(["Date", "Title"]); ws.append(["2025-05-21", "Sheet Title"])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    client.post("/media/spreadsheet", data={"file": (buf, "p.xlsx")},
                content_type="multipart/form-data")
    client.post("/media/mapping", json={
        "sheet_name": "Plan", "date_column": "Date", "youtube_title_column": "Title",
    })

    run_id = _init(client)
    file_id = _new_file(client, run_id)
    _complete_file(client, run_id, file_id)
    resp = client.post("/media/batch/run", json={
        "run_id": run_id, "dates": ["2025-05-21"], "platforms": ["youtube_video"],
        "files": {file_id: {"category": "youtube_video", "date": "2025-05-21"}},
        "overrides": {"2025-05-21": {"youtube_title": "Edited Title"}},
    })
    assert resp.status_code == 200
    job_id = resp.get_json()["job_id"]
    deadline = time.time() + 10
    while time.time() < deadline and not (upload_jobs.get_job(job_id) or {}).get("done"):
        time.sleep(0.05)
    assert captured["entries"]["2025-05-21"].youtube_title == "Edited Title"


def test_batch_overrides_apply_episode_title_and_vista_caption(client, monkeypatch):
    """The platform-tabs review widened the override schema to include
    episode_title (SimpleCast) and vista_caption (Vista Social). Both
    must round-trip through ``/media/batch/run`` so per-platform edits
    in the new tabbed review actually reach ``ReviewEntry`` and end
    up in the upload payload."""
    import io
    import openpyxl
    captured = {}
    monkeypatch.setattr(upload_jobs, "run_batch",
                        lambda **k: (captured.update(entries=k["entries_snapshot"]) or
                                     set(k["file_paths"].values())))
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Plan"
    ws.append(["Date"]); ws.append(["2025-05-21"])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    client.post("/media/spreadsheet", data={"file": (buf, "p.xlsx")},
                content_type="multipart/form-data")
    client.post("/media/mapping", json={"sheet_name": "Plan", "date_column": "Date"})
    run_id = _init(client)
    file_id = _new_file(client, run_id)
    _complete_file(client, run_id, file_id)
    resp = client.post("/media/batch/run", json={
        "run_id": run_id, "dates": ["2025-05-21"],
        "platforms": ["simplecast", "vista_social"],
        "files": {file_id: {"category": "podcast", "date": "2025-05-21"}},
        "overrides": {"2025-05-21": {
            "episode_title": "Pilot Episode",
            "vista_caption": "Custom caption ✨",
        }},
    })
    assert resp.status_code == 200
    job_id = resp.get_json()["job_id"]
    deadline = time.time() + 10
    while time.time() < deadline and not (upload_jobs.get_job(job_id) or {}).get("done"):
        time.sleep(0.05)
    entry = captured["entries"]["2025-05-21"]
    assert entry.episode_title == "Pilot Episode"
    assert entry.vista_caption == "Custom caption ✨"


def test_batch_run_agent_path_works_without_uploaded_files(client, monkeypatch):
    """Agent path: the browser shouldn't have to chunk-upload anything
    because the files already live on the agent's local disk (the
    agent resolves paths via its own ``agent.scan``). The previous
    flow chunk-uploaded anyway and deleted server-side immediately
    after dispatch — pure bandwidth waste.

    This test confirms the server accepts a batch_run dispatch with
    ``files: {}`` and real dates/platforms, builds entries with all
    path fields = None, and hands them to ``agent_dispatch.start``
    without a reassembly error. The agent will fill in paths from
    its scan when it processes the job_plan."""
    import io, openpyxl
    monkeypatch.setenv("HYBRID_AGENT_ENABLED", "true")

    # Stub agent_dispatch.start so we can inspect the entries snapshot
    # that's about to be sent to the agent.
    captured = {}

    def _fake_start(**kw):
        captured.update(kw)
        return "JOB-AGENT"

    from core import agent_dispatch
    monkeypatch.setattr(agent_dispatch, "start", _fake_start)
    monkeypatch.setattr(agent_dispatch, "register_job",
                        lambda **kw: None)

    # Prime a real sheet so build_entry has metadata to work with.
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Plan"
    ws.append(["Date", "Title"])
    ws.append(["2026-06-01", "Day 1"])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    client.post("/media/spreadsheet", data={"file": (buf, "plan.xlsx")},
                content_type="multipart/form-data")
    client.post("/media/mapping", json={
        "sheet_name": "Plan", "date_column": "Date",
        "youtube_title_column": "Title",
    })

    # Initialize a run (acquires the per-user lock) and immediately
    # dispatch on the agent path with NO file uploads.
    run_id = _init(client)
    resp = client.post("/media/batch/run?path=agent", json={
        "run_id": run_id,
        "dates": ["2026-06-01"],
        "platforms": ["youtube_video"],
        "files": {},  # ← key change: no upload required
    })
    assert resp.status_code == 200, resp.get_data(as_text=True)
    assert resp.get_json()["job_id"] == "JOB-AGENT"

    # The agent received a non-empty entries snapshot, but with no
    # server-side paths — those get resolved by the agent's scan.
    entries = captured.get("entries") or {}
    assert "2026-06-01" in entries
    entry = entries["2026-06-01"]
    # Title flowed through from the spreadsheet.
    assert entry.youtube_title == "Day 1"
    # Path fields are all None — the agent fills them in via scan.
    for f in ("youtube_video_path", "youtube_shorts_path",
              "podcast_path", "thumbnail_path"):
        assert getattr(entry, f) is None, f"{f} should be None on agent path"


def test_run_finish_releases_lock(client):
    run_id = _init(client)
    # Busy while active.
    assert client.post("/media/run/init", json={}).status_code == 409
    assert client.post("/media/run/finish", json={"run_id": run_id}).status_code == 200
    # Lock released → a new run can start.
    assert client.post("/media/run/init", json={}).status_code == 200


def test_sweep_orphans_removes_inactive_run(client, tmp_path):
    from blueprints import media
    run_id = _init(client)
    run_path = media._runs[run_id]["dir"].path
    import os
    assert os.path.isdir(run_path)
    # An orphan = a temp dir whose run id isn't in the active set.
    removed = ms.sweep_orphans(active_run_ids=set())
    assert removed >= 1
    assert not os.path.exists(run_path)
