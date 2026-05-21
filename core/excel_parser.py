"""
Excel (.xlsx) parser for upload metadata.

The user selects which sheet and which columns contain their data.
Column mapping is stored in config.yaml and editable in the Settings UI.
There is no required format — any sheet layout works as long as the user
maps the columns correctly.
"""

import logging
import os
import re
from datetime import datetime, date
from pathlib import Path
from typing import Iterator, Optional

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


_METADATA_FIELDS = (
    "shorts_title", "description", "youtube_title", "podcast_title",
    "transcript", "passage", "scripture", "episode_title", "prayer",
    "topic", "vista_caption",
)


def _empty_metadata() -> dict:
    out: dict = {fld: "" for fld in _METADATA_FIELDS}
    out["tags"] = []
    return out


def _normalize_headers(raw: list) -> list[str]:
    """Match pandas' header normalization so existing column mappings keep working.

    - Blank / None header cells become ``Unnamed: <0-indexed column>``.
    - Duplicate names get ``.1``, ``.2`` suffixes in order of appearance.
    """
    headers: list[str] = []
    for idx, value in enumerate(raw):
        if value is None or str(value).strip() == "":
            headers.append(f"Unnamed: {idx}")
        else:
            headers.append(str(value))

    seen: dict[str, int] = {}
    deduped: list[str] = []
    for name in headers:
        n = seen.get(name, 0)
        deduped.append(name if n == 0 else f"{name}.{n}")
        seen[name] = n + 1
    return deduped


def _cell_to_str(value) -> str:
    """Render a cell value the way pandas' ``fillna("").astype(str)`` would.

    Empty/None becomes "" — everything else gets ``str()``. Dates are kept
    as ``datetime`` here and stringified by callers when needed; for raw
    string cells this matches the prior pandas behavior.
    """
    if value is None:
        return ""
    return str(value)


def _iter_rows(xlsx_path: str, sheet_name: Optional[str] = None,
               max_rows: Optional[int] = None) -> Iterator[dict]:
    """Yield each data row of *sheet_name* as ``{header: value}``.

    Uses ``read_only=True`` so multi-megabyte workbooks don't load fully
    into memory. ``data_only=True`` returns formula results, not formulas.
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        row_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(row_iter)
        except StopIteration:
            return
        headers = _normalize_headers(list(header_row))

        for i, row in enumerate(row_iter):
            if max_rows is not None and i >= max_rows:
                break
            # Pad/truncate so zip() lines up even when openpyxl trims
            # trailing empty cells on a given row.
            row_padded = list(row) + [None] * (len(headers) - len(row))
            yield dict(zip(headers, row_padded[:len(headers)]))
    finally:
        wb.close()


def get_sheet_names(xlsx_path: str) -> list[str]:
    """Return list of sheet names in the workbook."""
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    except Exception as e:
        logger.error("Could not open Excel file %s: %s", xlsx_path, e, exc_info=True)
        return []


def get_column_names(xlsx_path: str, sheet_name: str) -> list[str]:
    """Return column names (header row) for a given sheet."""
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                return _normalize_headers(list(row))
            return []
        finally:
            wb.close()
    except Exception as e:
        logger.error("Could not read columns from sheet %s: %s", sheet_name, e, exc_info=True)
        return []


def get_sheet_preview(xlsx_path: str, sheet_name: str, num_rows: int = 5) -> list[dict]:
    """Return first N rows as list of dicts for preview in UI."""
    try:
        out: list[dict] = []
        for row in _iter_rows(xlsx_path, sheet_name, max_rows=num_rows):
            out.append({k: _cell_to_str(v) for k, v in row.items()})
        return out
    except Exception as e:
        logger.error("Could not preview sheet %s: %s", sheet_name, e, exc_info=True)
        return []


class ExcelParser:
    def __init__(self, config: dict):
        self.xlsx_path = config.get("sharepoint_docx", "")  # reusing same key
        self.mapping = config.get("excel_mapping", {})
        self._cache: Optional[dict] = None
        self._cache_mtime: Optional[float] = None
        self.last_error: str = ""

    def _file_mtime(self) -> Optional[float]:
        try:
            return os.path.getmtime(self.xlsx_path) if self.xlsx_path else None
        except OSError:
            return None

    def _load(self) -> dict[str, dict]:
        current_mtime = self._file_mtime()
        if self._cache is not None and current_mtime == self._cache_mtime:
            return self._cache

        try:
            exists = bool(self.xlsx_path) and Path(self.xlsx_path).exists()
        except OSError as e:
            logger.warning("Excel path check failed for %s: %s", self.xlsx_path, e)
            self.last_error = f"Excel path unreachable: {e}"
            self._cache = {}
            return self._cache
        if not exists:
            logger.warning("Excel file not found: %s", self.xlsx_path)
            self.last_error = f"Excel file not found: {self.xlsx_path}" if self.xlsx_path else "No Excel file configured"
            self._cache = {}
            return self._cache

        sheet = self.mapping.get("sheet_name")
        date_col = self.mapping.get("date_column")

        if not sheet or not date_col:
            logger.warning("Excel mapping not configured — set sheet and date column in Settings")
            self.last_error = "Excel mapping not configured (sheet + date column required)"
            self._cache = {}
            return self._cache

        try:
            rows = list(_iter_rows(self.xlsx_path, sheet))
        except Exception as e:
            logger.error("Failed to read Excel file: %s", e)
            self.last_error = f"Failed to read Excel file: {e}"
            self._cache = {}
            return self._cache

        self.last_error = ""

        result: dict = {}
        skipped_dates = 0
        for row in rows:
            raw_date = row.get(date_col, "")
            dt = self._parse_date(raw_date)
            if dt is None:
                if raw_date and str(raw_date).strip():
                    skipped_dates += 1
                    logger.debug("Excel: skipped row with unparseable date %r", raw_date)
                continue

            def col(key, fallback=""):
                col_name = self.mapping.get(key)
                if col_name and col_name in row:
                    val = row[col_name]
                    return _cell_to_str(val).strip()
                return fallback

            entry = _empty_metadata()
            entry.update({
                "shorts_title":   col("shorts_title_column"),
                "description":    col("description_column"),
                "tags":           [t.strip() for t in col("tags_column").split(",") if t.strip()],
                "youtube_title":  col("youtube_title_column"),
                "podcast_title":  col("podcast_title_column"),
                "transcript":     col("transcript_column"),
                "passage":        col("passage_column"),
                "scripture":      col("scripture_column"),
                "episode_title":  col("episode_title_column"),
                "prayer":         col("prayer_column"),
                "topic":          col("topic_column"),
                "vista_caption":  col("vista_caption_column"),
            })
            result[dt.isoformat()] = entry

        if skipped_dates:
            logger.info("Excel: skipped %d row(s) with unparseable dates", skipped_dates)
        self._cache = result
        self._cache_mtime = current_mtime
        return result

    def _parse_date(self, value) -> Optional[date]:
        if isinstance(value, (datetime, date)):
            return value.date() if isinstance(value, datetime) else value
        if not value or str(value).strip() == "":
            return None
        text = str(value).strip()
        formats = [
            "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y",
            "%B %d, %Y", "%b %d, %Y",
            "%m/%d/%y", "%d/%m/%y",
        ]
        for fmt in formats:
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        digits = re.sub(r"\D", "", text)
        if len(digits) == 6:
            try:
                return datetime.strptime(
                    f"20{digits[0:2]}-{digits[2:4]}-{digits[4:6]}", "%Y-%m-%d"
                ).date()
            except ValueError:
                pass
        return None

    def get_metadata(self) -> dict:
        return self._load()

    def get_metadata_for_date(self, date_str: str) -> dict:
        data = self._load()
        return data.get(date_str, _empty_metadata())

    def invalidate_cache(self):
        self._cache = None
        self._cache_mtime = None
